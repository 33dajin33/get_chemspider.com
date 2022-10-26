# @Time : ${DATE} ${TIME}
# @Author : CandyT
# @GitHub : https://github.com/33dajin33
import sys
import os
import requests
import re
from bs4 import BeautifulSoup
import csv
import pandas
from openpyxl import load_workbook
import time

header = {
    "X-Requested-With":"XMLHttpRequest",
    "Cache-Control":"no-cache",
    "Accept":"*/*",
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36 Edg/106.0.1370.52",
    "Accept-Encoding":"gzip, deflate",
    "Accept-Language":"zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
    "Connection":"close"
    }
#完成
def write_data(line,id,value1,value2):
    wb = load_workbook("F://pycode//test//spiderfield//read.xlsx")
    ws = wb.active
    ws.cell(line+2, 1, id)
    ws.cell(line+2, 2, value1)
    ws.cell(line+2, 3, value2)
    wb.save("F://pycode//test//spiderfield//read.xlsx")
    


#完成
def get_url(hmdb_list,header,id):
    url = "http://www.chemspider.com/ibcontent.ashx?csid="+id+"&type=ds&ds_types="
    res = requests.get(url=url,headers=header)
    soup = BeautifulSoup(res.text,"html.parser")
    soup = soup.select("#DS_ALL a")
    for i in soup:
        if "HMDB" in i["href"]:
            hmdb_list.append(i["href"])


def get_content(new_Adducts,hmdb_list,header,first,second,Adducts,value):
    #res = requests.get(hmdb_list[0],headers=header)
    for url in hmdb_list:
        res = requests.get(url=url,headers=header)
        soup = BeautifulSoup(res.text,"html.parser")
        print("正在访问"+url)
        soup = soup.find("table",class_="table table-bordered ccs")
        tr = soup.tbody.find_all("tr")
        #循环每一个tr,每一个tr下的td是一个数组col
        for i in tr:
            td = i.find_all("td")
            col = [j.text.strip() for j in td]
            first.append(col[1])
            second.append(col[2])
            time.sleep(3)
    length = len(first)
    for i in new_Adducts:
        for j in range(length):
            if i in first[j]:
                Adducts.append(first[j])
                value.append(second[j])
'''
            for lii in li:
                if lii in col[1]:
                    first.append(col[1])
                    second.append(col[2])
        
        if len(first) == len(second):
            length = len(first)
            for i in new_Adducts:
                for j in range(length):
                    if i in first[j]:
                        Adducts.append(first[j])
                        value.append(second[j])
                        time.sleep(3)

        else:
            print("ERROR!!!!")
'''        
                

#完成
def get_Adducts_id(New_adducts,line):
    df = pandas.read_csv("F://pycode//test//spiderfield//read.csv")
    Adducts = df.iloc[line]["Adducts"]
    Adducts = Adducts.split(", ")
    for i in Adducts:
        New_adducts.append("["+i+"]")
    id = df.iloc[line]["id"]
    return str(id)






if __name__ =="__main__":

    """
    hmdb_list = []
    #get_content(hmdb_list,header)
    id = "24769296"
    get_url(header,id)
    print(hmdb_list)
    """

    for line in range(9):
        print("第"+str(line+1)+"行数据")
        #获取Adducts、id值
        new_Adducts = []
        id = get_Adducts_id(new_Adducts,line)
        #print("HMDB的id "+id)
        print(new_Adducts)
        #获取hmdb链接
        hmdb_list = []
        first = []
        second = []
        Adducts = []
        value = []
        get_url(hmdb_list,header,id)
        if len(hmdb_list) == 0:
            value1 = "无HMDB值"
            value2 = "无Adducts值"
            write_data(line,id,value1,value2)    
        else:
            #print(hmdb_list)
            get_content(new_Adducts,hmdb_list,header,first,second,Adducts,value)
            write_data(line,id,",".join(value),",".join(Adducts))
        print("第"+str(line+1)+"行爬取结束")
        time.sleep(3)

